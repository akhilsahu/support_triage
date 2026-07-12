"""
XlsxParser — openpyxl for .xlsx / .xls.
Each sheet becomes a section. Header row is preserved as column labels.
"""

from __future__ import annotations

import io
from pathlib import Path
from typing import List

import structlog

from app.rag.document_parser import ParsedDocument, ParsedPage
from app.orchestra.ai.ingestion.core.base import BaseParser

logger = structlog.get_logger()


class XlsxParser(BaseParser):
    extensions = [".xlsx", ".xls"]

    def parse(self, raw: bytes, filename: str) -> ParsedDocument:
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

        # openpyxl reads .xlsx natively; for .xls we convert via xlrd if available
        ext = Path(filename).suffix.lower()
        if ext == ".xls":
            raw = self._xls_to_xlsx(raw, filename)

        wb    = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
        pages: List[ParsedPage] = []

        for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
            ws      = wb[sheet_name]
            rows    = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # First non-empty row = header
            header = next(
                ([str(c) if c is not None else "" for c in row] for row in rows if any(c is not None for c in row)),
                None,
            )
            if not header:
                continue

            lines: List[str] = [" | ".join(header)]
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    lines.append(" | ".join(cells))

            text = f"Sheet: {sheet_name}\n" + "\n".join(lines)
            pages.append(ParsedPage(page=sheet_idx, text=text, section=sheet_name, is_table=True))

        meta = {"sheet_count": len(wb.sheetnames), "sheets": wb.sheetnames}
        logger.info("ingestion.xlsx", filename=filename, sheets=len(pages))
        return ParsedDocument(filename=filename, extension=ext, pages=pages, metadata=meta)

    def _xls_to_xlsx(self, raw: bytes, filename: str) -> bytes:
        """Convert legacy .xls to .xlsx bytes via xlrd + openpyxl."""
        try:
            import xlrd
            import openpyxl
            import io as _io

            wb_old = xlrd.open_workbook(file_contents=raw)
            wb_new = openpyxl.Workbook()
            wb_new.remove(wb_new.active)

            for sheet in wb_old.sheets():
                ws = wb_new.create_sheet(title=sheet.name)
                for row in range(sheet.nrows):
                    for col in range(sheet.ncols):
                        ws.cell(row=row + 1, column=col + 1, value=sheet.cell_value(row, col))

            buf = _io.BytesIO()
            wb_new.save(buf)
            logger.info("ingestion.xls.converted", filename=filename)
            return buf.getvalue()
        except ImportError:
            raise RuntimeError(
                "xlrd not installed for .xls conversion. Run: pip install xlrd openpyxl"
            )
